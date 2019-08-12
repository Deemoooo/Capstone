import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.descriptors import Bool, Alias

#from pyxlsb import open_workbook

# bank statement data
# bankdata = pd.read_excel('bankstatement.xlsm')
# billdata = pd.read_excel('billstatement.xlsx')
# loandata = pd.read_excel('loanstatement.xlsx')

# CAT workbook
#wb2 = open_workbook("Obligor 1 - CAT spreadsheet.xlsb")

#extracting information (someone helpppp)
#with open_workbook('Obligor 1 - CAT spreadsheet.xlsb') as wbcat:
#    for sheetname in wbcat.sheets:
#        with wbcat.get_sheet(sheetname) as sheet:
#            for row in sheet.rows():
#                values = [r.v for r in row]  # retrieving content
#                csv_line = ','.join(values)  # or do your thing
#                print(csv_line)

def processor():
    # create workbook
    wb = Workbook()
    wb = load_workbook('static/output/output.xlsm', keep_vba=True)

    # create worksheets
    # bankstatement = wb.active
    # bankstatement.title = 'Bank Statement'
    # billstatement = wb.create_sheet("Bill Statement")
    # loanrepayment = wb.create_sheet("Loan Repayment Schedule")
    # incomestatement = wb.create_sheet("Income Statement")
    # balancesheet = wb.create_sheet("Balance Sheet")
    #calculatedcf = wb.create_sheet("Calculated CF")
    usermetrics = wb.create_sheet("User Metric")
    summary = wb.create_sheet("Summary")

    # add bank statement data in bankstatement worksheet
    # for r in dataframe_to_rows(bankdata, index=False, header=True):
    #     bankstatement.append(r)
    # bankstatement.delete_rows(1)
    #
    # # add bill statement data in billstatement worksheet
    # for r in dataframe_to_rows(billdata, index=False, header=True):
    #     billstatement.append(r)
    #
    # # add loan repayment data in loanrepayment worksheet
    # for r in dataframe_to_rows(loandata, index=False, header=True):
    #     loanrepayment.append(r)

    # add balance sheet from CAT to balancesheet worksheet
        # need to figure this out
    #balancesheet = wb2.get_sheet("Balance Sheet")

    # assign fixed variables to user metrics


    # User Input
    # create drop down list for user input

    ## Type of Client
    clientdv = DataValidation(type="list", formula1='"new-to-bank,existing"', allow_blank=False)
    clientdv.prompt = 'Please select from the list'
    clientdv.promptTitle = 'Client List Selection'

    ## Payment status
    paymentdv = DataValidation(type="list", formula1='"punctual,not punctual"', allow_blank=False)
    paymentdv.prompt = 'Please select from the list'
    paymentdv.promptTitle = 'Payment List Selection'

    ## Operating account
    oadv = DataValidation(type="list", formula1='"reasonable,not reasonable"', allow_blank=False)
    oadv.prompt = 'Please select from the list'
    oadv.promptTitle = 'Behavior List Selection'

    ## Liquidity status
    liquiditydv = DataValidation(type="list", formula1='"low,medium,high"', allow_blank=False)
    liquiditydv.prompt = 'Please select from the list'
    liquiditydv.promptTitle = 'Liquidity Level List Selection'

    ## TU Information
    tudv = DataValidation(type="list", formula1='"found,not found"', allow_blank=False)
    tudv.prompt = 'Please select from the list'
    tudv.promptTitle = 'Behavior List Selection'

    # Summary
    # color scheme
    greyFill = PatternFill(start_color='757171',
                       end_color='757171',
                       fill_type='solid')

    whiteFill = PatternFill(start_color='d0cece',
                       end_color='d0cece',
                       fill_type='solid')

    # text style and color
    whiteheadingstyle = Font(size = 14, name = 'Segoe UI', color = 'ffffff', underline='single')
    whitebodystyle = Font(size = 11, name = 'Segoe UI', color = 'ffffff')
    greybodystyle = Font(size = 11, name = 'Segoe UI', color = '757171')

    # width
    width = 2.8

    # formatting
    # change background color
    # No of written Rows in sheet
    # 1,048,576 rows by 16,384 columns

    # No of written Rows in sheet
    r = 500
    # No of written Columns in sheet
    c = 100

    # mass format color, font, filler lines, result lines and account background sentences in excel
    filler_lines_in_grey_square = ["Please fill in the blanks accordingly via the dropdown options", \
                                   "(The inputs will be used within the sentences on the left)", \
                                   "Description", \
                                   "Type of client", \
                                   "Type of business (SIC)", \
                                   "Type of business (NAICS)", \
                                   "Number of Installment Loans", \
                                   "Payment Status", \
                                   "Operating Account", \
                                   "Liquidity Level", \
                                   "Negative TU Information"]

    result_lines_in_grey_square = [clientdv, \
                                   '=VLOOKUP(\'Income Statement\'!$E$2, SIC!$A$4:$B$1008,2,FALSE)', \
                                   '=VLOOKUP(\'Income Statement\'!$E$3, NAICS!$B$4:$C$2231,2,FALSE)', \
                                   "Fill in here", \
                                   paymentdv, \
                                   oadv, \
                                   liquiditydv, \
                                   tudv]

    account_background_sentences = ['=CONCATENATE("This is a ",\'User Input\'!B2," client of Citi. At present, client also incorporated bank accounts in Citi for ",$G$7," activities.")', \
                                    '=CONCATENATE("The company engages in ",$G$6," business, sales proceeds collection was captured in bank account.")', \
                                    '=CONCATENATE("With reference to ",\'User Input\'!B3," bank statement, about ",DOLLAR(Final!D27)," credit transaction was recorded. The annualised sales proceeds was estimated to be ",DOLLAR(Final!D28),". If subject have more than 2 banks, state the transaction figures and periods.")', \
                                    'Such figure could serve as an estimate for annual sales of subject borrower. ', \
                                    '=CONCATENATE("As at ",TEXT(MONTH(\'Bank Statement\'!A44),"mmm")," ",YEAR(\'Bank Statement\'!A44),", client is keeping about ",DOLLAR(\'Bank Statement\'!E43)," bank balance in Citibank. ")',\
                                    '=CONCATENATE("No negative record could be found in TU and CCRA on ", TEXT(MONTH(\'Bank Statement\'!A44),"mmm")," ",YEAR(\'Bank Statement\'!A44),".")', \
                                    '=CONCATENATE(G8," instalment loans was respectively booked for BankA (Loan Amount, number of years) e.g. HKD 5MM for 10 years, and it is revealed by TU repayment of these instalment has been ",G9,".")', \
                                    '=CONCATENATE("As per bank statement from Citibank provided, it is observed that Subject received ",DOLLAR(Final!D28)," throughput (Annualized: $$$)")', \
                                    '==CONCATENATE("AO confirmed that aggregated operating account for $$$ is considered ",G10)', \
                                    '=CONCATENATE("Average free cash flow is around ",DOLLAR(Final!E27)," in bank statement. ", \'User Input\'!B7," liquidity is observed.")', \
                                    '=CONCATENATE("Negative information is ",G12, " on TU.")']

    types_of_data_validation = [clientdv,oadv,liquiditydv,tudv]

    for i in range(1, r+1):
        for j in range(1, c+1):
            summary.cell(row=i, column=j).fill = greyFill
            summary.cell(row=i, column=j).font = whitebodystyle
            if j == 4:
                if i == 2:
                    # heading style
                    summary.cell(row=i, column=j).font = whiteheadingstyle
                    # add filler lines
                    summary.cell(row=i, column=j).value = 'ACCOUNT BACKGROUND (For your designated usage)'
                elif i > 3 and i < 15:
                    # add account background sentences
                    summary.cell(row=i, column=j).value = account_background_sentences[i-4]
                    summary.cell(row=i, column=j).alignment = Alignment(wrap_text=Alias('wrapText'), wrapText = Bool(False))

            # labelling the sentences
            if i>3 and i<15 and j==2:
                summary.cell(row=i, column=j).value = i-3

    # create grey square in excel, with designated font
    for i in range(2,14):
        for j in range(5,9):
            if i == 4 and j == 6:
                summary.cell(row=i, column=j).fill = greyFill
                summary.cell(row=i, column=j).font = whitebodystyle
            elif i == 4 and j == 7:
                summary.cell(row=i, column=j).fill = greyFill
                summary.cell(row=i, column=j).font = whitebodystyle
            else:
                summary.cell(row=i, column=j).fill = whiteFill
                summary.cell(row=i, column=j).font = greybodystyle
            # add filler lines and data validation
            if j == 6 and i < 13:
                summary.cell(row=i, column=j).value = filler_lines_in_grey_square[i-2]
                if i > 4:
                    if type(result_lines_in_grey_square[i-5]) == str:
                        summary.cell(row=i, column=j+1).value = result_lines_in_grey_square[i-5]
                        summary.cell(row=i, column=j+1).alignment = Alignment(wrap_text=Alias('wrapText'), wrapText = Bool(False))
                    else:
                        summary.add_data_validation(result_lines_in_grey_square[i-5])
                        result_lines_in_grey_square[i-5].add(summary.cell(row=i, column=j+1))
                        summary.cell(row=i, column=j+1).value = "Select from drop-down"

            if j == 7 and i == 3:
                summary.cell(row=i, column=j).value = "To be filled"

    # set the width of the column
    specialcolumns = ['A','B','C','E','H']

    for i in specialcolumns:
        summary.column_dimensions[i].width = width
    summary.column_dimensions['D'].width = 105
    summary.column_dimensions['F'].width = 29
    summary.column_dimensions['G'].width = 49.64

    # adjust zoom level
    summary.sheet_view.zoomScale = 85

    # Save the file
    wb.save("static/output/output.xlsm")
    
